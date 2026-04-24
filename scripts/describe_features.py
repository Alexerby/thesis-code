"""
Descriptive analysis of extracted features for each confirmed spoofing event.

For each event in events.json, splits feature records into in-window and
out-of-window groups using ts_recv, then plots KDE distributions side by side
for every feature.

Usage:
    python scripts/describe_features.py --features output/MULN_20221025/FEATURES.csv
    python scripts/describe_features.py  # runs all events found in events.json
"""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime, timezone

import matplotlib.pyplot as plt
import openpyxl.styles as xlstyles
import pandas as pd

EVENTS_FILE = "events.json"
BINARY = "./build/thesis"
FEATURES = [
    "delta_t",
    "induced_imbalance",
    "volume_ahead",
    "relative_size",
    "price_distance_ticks",
    "spread_bps",
]
FEATURE_LABELS = {
    "delta_t":              "Order lifetime (ns)",
    "induced_imbalance":    "Induced OBI change",
    "volume_ahead":         "Volume ahead at add",
    "relative_size":        "Relative size",
    "price_distance_ticks": "Price distance (ticks)",
    "spread_bps":           "Spread (bps)",
}


def window_to_ns(date_str, time_str):
    """Convert a date (YYYY-MM-DD) + time (HH:MM:SS) in ET to nanoseconds since epoch."""
    import zoneinfo

    dt_et = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
    dt_et = dt_et.replace(tzinfo=zoneinfo.ZoneInfo("America/New_York"))
    return int(dt_et.timestamp() * 1e9)


CANCEL_TYPE_LABELS = {0: "Pure", 1: "Fill"}


def load_features(path):
    """Read a FEATURES.csv, add human-readable columns, and save back in place."""
    df = pd.read_csv(path)
    if "ts_recv_et" not in df.columns:
        df.insert(
            df.columns.get_loc("ts_recv") + 1, "ts_recv_et", ns_to_et_str(df["ts_recv"])
        )
    if "cancel_type" in df.columns and "cancel_type_label" not in df.columns:
        df.insert(
            df.columns.get_loc("cancel_type") + 1,
            "cancel_type_label",
            df["cancel_type"].map(CANCEL_TYPE_LABELS).fillna("Unknown"),
        )
    df.to_csv(path, index=False)
    return df


def ns_to_et_str(ns_series):
    """Convert a Series of UTC nanoseconds to ET datetime strings."""
    import zoneinfo

    return (
        pd.to_datetime(ns_series, unit="ns", utc=True)
        .dt.tz_convert(zoneinfo.ZoneInfo("America/New_York"))
        .dt.strftime("%Y-%m-%d %H:%M:%S.%f ET")
    )


def cohens_d(s_in, s_out):
    n_in, n_out = len(s_in), len(s_out)
    if n_in < 2 or n_out < 2:
        return float("nan")
    pooled_std = (
        ((n_in - 1) * s_in.std() ** 2 + (n_out - 1) * s_out.std() ** 2)
        / (n_in + n_out - 2)
    ) ** 0.5
    return (s_in.mean() - s_out.mean()) / pooled_std if pooled_std > 0 else float("nan")


def compute_descriptives(in_window, out_window):
    rows = []
    for feat in FEATURES:
        if feat not in in_window.columns:
            continue
        s_in = in_window[feat].dropna()
        s_out = out_window[feat].dropna()
        d = cohens_d(s_in, s_out)
        rows.append(
            {
                "Feature": FEATURE_LABELS.get(feat, feat),
                "In N": len(s_in),
                "In Mean": s_in.mean(),
                "In Median": s_in.median(),
                "In Std": s_in.std(),
                "Out N": len(s_out),
                "Out Mean": s_out.mean(),
                "Out Median": s_out.median(),
                "Out Std": s_out.std(),
                "Cohen's d": d,
            }
        )
    return pd.DataFrame(rows)


def describe_event(df, event, output_dir):
    df = df.copy()
    df["ts_recv_et"] = ns_to_et_str(df["ts_recv"])
    if "cancel_type" in df.columns:
        df["cancel_type_label"] = (
            df["cancel_type"].map(CANCEL_TYPE_LABELS).fillna("Unknown")
        )

    date = event["date"]
    win_start = window_to_ns(date, event["window_start"])
    win_end = window_to_ns(date, event["window_end"])

    in_window = df[(df["ts_recv"] >= win_start) & (df["ts_recv"] <= win_end)]
    out_window = df[(df["ts_recv"] < win_start) | (df["ts_recv"] > win_end)]

    ts_min = df["ts_recv_et"].iloc[0] if len(df) else "n/a"
    ts_max = df["ts_recv_et"].iloc[-1] if len(df) else "n/a"

    print(f"\n{'=' * 60}")
    print(f"  {date}  |  window {event['window_start']} – {event['window_end']} ET")
    print(f"  Data range    : {ts_min}")
    print(f"                  {ts_max}")
    print(f"  Total records : {len(df):,}")
    print(f"  In window     : {len(in_window):,}")
    print(f"  Out of window : {len(out_window):,}")
    print(f"{'=' * 60}")

    # --- Summary table ---
    print(f"\n{'Feature':<25} {'In mean':>12} {'Out mean':>12} {'Cohens d':>10}")
    print("-" * 62)
    for feat in FEATURES:
        if feat not in df.columns:
            continue
        s_in = in_window[feat].dropna()
        s_out = out_window[feat].dropna()
        d = cohens_d(s_in, s_out)
        print(f"{feat:<25} {s_in.mean():>12.4f} {s_out.mean():>12.4f} {d:>10.3f}")

    if "cancel_type_label" in df.columns:
        for label, grp in [("in window    ", in_window), ("out of window", out_window)]:
            counts = grp["cancel_type_label"].value_counts()
            print(
                f"  cancel_type {label}: Pure={counts.get('Pure', 0):,}  Fill={counts.get('Fill', 0):,}"
            )

    # --- Boxplots ---
    n_features = len([f for f in FEATURES if f in df.columns])
    ncols = 3
    nrows = (n_features + ncols - 1) // ncols

    fig, axes = plt.subplots(nrows, ncols, figsize=(15, 4 * nrows))
    fig.suptitle(
        f"Feature distributions — MULN {date}\n"
        f"Window {event['window_start']}–{event['window_end']} ET  "
        f"(in={len(in_window):,}, out={len(out_window):,})",
        fontsize=13,
        fontweight="bold",
    )

    ax_flat = axes.flat
    for feat in FEATURES:
        if feat not in df.columns:
            continue
        ax = next(ax_flat)

        combined = pd.concat([in_window[feat], out_window[feat]])
        lo, hi = combined.quantile(0.01), combined.quantile(0.99)

        bp = ax.boxplot(
            [in_window[feat].clip(lo, hi), out_window[feat].clip(lo, hi)],
            patch_artist=True,
            widths=0.5,
            showfliers=False,
            medianprops=dict(color="black", linewidth=2),
        )
        bp["boxes"][0].set_facecolor("crimson")
        bp["boxes"][0].set_alpha(0.6)
        bp["boxes"][1].set_facecolor("steelblue")
        bp["boxes"][1].set_alpha(0.6)

        ax.set_xticks([1, 2])
        ax.set_xticklabels(["In window", "Out of window"], fontsize=9)
        ax.set_title(FEATURE_LABELS.get(feat, feat), fontsize=10)
        ax.set_ylabel("")
        ax.grid(axis="y", alpha=0.3)

    for ax in ax_flat:
        ax.set_visible(False)

    plt.tight_layout()
    compact = date.replace("-", "")
    out_path = os.path.join(output_dir, f"DISTRIBUTIONS_{compact}.png")
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"\n  Plot saved: {out_path}")

    return compute_descriptives(in_window, out_window)


def _color_d(d):
    """Return a hex fill color for a Cohen's d value. Saturates at |d| = 2."""
    try:
        d = float(d)
        intensity = min(int(abs(d) / 2 * 200 + 55), 255)
        return (f"00{255 - intensity:02X}00") if d > 0 else (f"{255 - intensity:02X}0000")
    except (TypeError, ValueError):
        return None


def _write_consolidated_excel(sheets, xl_path):
    """
    One sheet, features as rows, events as column groups.
    Each group: In Median | Out Median | Cohen's d
    This makes it easy to scan each feature across all events at once.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    SUB_COLS = ["In Median", "Out Median", "Cohen's d"]
    dates    = list(sheets.keys())
    features = [FEATURE_LABELS.get(f, f) for f in FEATURES]
    n_sub    = len(SUB_COLS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descriptives"

    bold      = xlstyles.Font(bold=True)
    center    = xlstyles.Alignment(horizontal="center")
    thin_side = xlstyles.Side(style="thin")

    # ── Row 1: "Feature" + date headers (merged across sub-columns) ──
    ws.cell(1, 1, "Feature").font = bold
    for i, date in enumerate(dates):
        col = 2 + i * n_sub
        cell = ws.cell(1, col, date)
        cell.font      = bold
        cell.alignment = center
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1,   end_column=col + n_sub - 1)

    # ── Row 2: sub-column headers ──
    ws.cell(2, 1, "")
    for i in range(len(dates)):
        for j, sub in enumerate(SUB_COLS):
            cell = ws.cell(2, 2 + i * n_sub + j, sub)
            cell.font      = bold
            cell.alignment = center

    # ── Data rows ──
    for row_i, feat_label in enumerate(features):
        ws.cell(row_i + 3, 1, feat_label)
        for col_i, date in enumerate(dates):
            df = sheets[date]
            match = df[df["Feature"] == feat_label]
            if match.empty:
                continue
            row_data = match.iloc[0]
            for sub_i, sub in enumerate(SUB_COLS):
                col = 2 + col_i * n_sub + sub_i
                val  = row_data.get(sub, float("nan"))
                cell = ws.cell(row_i + 3, col, float(val) if val == val else None)
                if sub == "Cohen's d":
                    color = _color_d(val)
                    if color:
                        cell.fill = xlstyles.PatternFill("solid", fgColor=color)
                    if val == val and abs(float(val)) > 0.8:
                        cell.font = bold
                cell.alignment = xlstyles.Alignment(horizontal="right")

    # ── Column widths ──
    ws.column_dimensions["A"].width = 26
    for i in range(len(dates)):
        for j in range(n_sub):
            letter = get_column_letter(2 + i * n_sub + j)
            ws.column_dimensions[letter].width = 13

    # ── Freeze feature column and header rows ──
    ws.freeze_panes = "B3"

    wb.save(xl_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument(
        "--features", help="Path to a single FEATURES.csv (overrides events.json loop)"
    )
    p.add_argument("--output-dir", default="output", help="Directory for saved plots")
    args = p.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    with open(EVENTS_FILE) as f:
        events_config = json.load(f)

    if not os.path.isfile(BINARY) or not os.access(BINARY, os.X_OK):
        sys.exit(f"Binary not found at {BINARY} — run cmake --build build first.")

    sheets = {}  # date -> descriptives DataFrame

    if args.features:
        df = load_features(args.features)
        basename = os.path.basename(args.features)
        matched = next(
            (
                e
                for e in events_config["events"]
                if e["date"].replace("-", "") in basename
            ),
            events_config["events"][0],
        )
        desc = describe_event(df, matched, args.output_dir)
        sheets[matched["date"]] = desc
    else:
        for event in events_config["events"]:
            compact = event["date"].replace("-", "")
            event_dir = f"output/MULN_{compact}"
            csv_path = f"{event_dir}/FEATURES.csv"
            dbn_path = event["dbn"]["event"]

            os.makedirs(event_dir, exist_ok=True)

            if not os.path.exists(csv_path):
                if not os.path.exists(dbn_path):
                    print(f"  Skipping {event['date']} — DBN not found: {dbn_path}")
                    continue
                print(f"  Extracting features: {dbn_path}")
                subprocess.run(
                    [
                        BINARY,
                        "extract-features",
                        dbn_path,
                        "--ticker",
                        "MULN",
                        "--output",
                        csv_path,
                    ],
                    check=True,
                )

            df = load_features(csv_path)
            desc = describe_event(df, event, event_dir)
            sheets[event["date"]] = desc

    if sheets:
        xl_path = os.path.join(args.output_dir, "DESCRIPTIVES.xlsx")
        _write_consolidated_excel(sheets, xl_path)
        print(f"\nDescriptives written to {xl_path}")


if __name__ == "__main__":
    main()

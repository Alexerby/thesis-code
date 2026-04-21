"""
Build a structured Excel workbook from output/comparison.csv.

Sheet 1 — SNR by Tier   : pivoted SNR table (tiers as rows, events as columns)
                          with conditional colour scale and average column.
                          This is the thesis story sheet.
Sheet 2 — Full Metrics  : all metrics for every date/tier combination.

Usage:
    python scripts/build_comparison_sheet.py \
        --input  output/comparison.csv \
        --output output/comparison.xlsx
"""

import argparse
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


TIER_LABELS = {
    1: "T1 — Baiting signature\n(rel_size, imbalance)",
    2: "T2 — + Cancellation timing\n(+ delta_t)",
    3: "T3 — + Order book context\n(+ dist, vol_ahead)",
    4: "T4 — Full model\n(+ spread_bps)",
    5: "T5 — SHAP-informed\n(rel_size, imbalance, dist)",
}

DATE_LABELS = {
    "2022-10-25": "Oct 25 2022\n14:26–14:28",
    "2022-12-15": "Dec 15 2022\n13:25–13:27",
    "2023-06-06": "Jun 06 2023\n15:50–15:52",
    "2023-08-17": "Aug 17 2023\n15:53–15:55",
}

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
AVG_FILL     = PatternFill("solid", fgColor="D6E4F0")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
LIGHT_FILL   = PatternFill("solid", fgColor="F2F7FB")

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT   = Font(bold=True, size=10)
BODY_FONT    = Font(size=10)
AVG_FONT     = Font(bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _pivot_sheet(wb, df, metric, sheet_name, title_suffix, col_label, avg_label, colour_rule):
    """Generic helper that builds a pivoted tier×event sheet for any metric."""
    ws = wb.create_sheet(sheet_name)
    dates  = sorted(df["Date"].unique())
    tiers  = sorted(df["Tier"].unique())
    pivot  = df.pivot(index="Tier", columns="Date", values=metric).reindex(index=tiers, columns=dates)

    ws.column_dimensions["A"].width = 34
    ws.row_dimensions[1].height = 40

    ws["A1"] = "Tier"
    ws["A1"].font = HEADER_FONT; ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER; ws["A1"].border = thin_border()

    for col_i, date in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=col_i, value=DATE_LABELS.get(date, date))
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = 20

    avg_col = len(dates) + 2
    avg_cell = ws.cell(row=1, column=avg_col, value=avg_label)
    avg_cell.font = HEADER_FONT; avg_cell.fill = SUBHEAD_FILL
    avg_cell.alignment = CENTER; avg_cell.border = thin_border()
    ws.column_dimensions[get_column_letter(avg_col)].width = 16

    for row_i, tier in enumerate(tiers, start=2):
        ws.row_dimensions[row_i].height = 38
        lc = ws.cell(row=row_i, column=1, value=TIER_LABELS.get(tier, f"T{tier}"))
        lc.font = LABEL_FONT
        lc.fill = LIGHT_FILL if row_i % 2 == 0 else WHITE_FILL
        lc.alignment = LEFT; lc.border = thin_border()

        row_vals = []
        for col_i, date in enumerate(dates, start=2):
            val = pivot.loc[tier, date] if date in pivot.columns else None
            try:
                val = round(float(val), 3)
            except (TypeError, ValueError):
                val = None
            cell = ws.cell(row=row_i, column=col_i, value=val if val is not None else "N/A")
            cell.font = BODY_FONT
            cell.fill = LIGHT_FILL if row_i % 2 == 0 else WHITE_FILL
            cell.alignment = CENTER; cell.border = thin_border()
            if val is not None:
                row_vals.append(val)

        avg_val = round(sum(row_vals) / len(row_vals), 3) if row_vals else "N/A"
        ac = ws.cell(row=row_i, column=avg_col, value=avg_val)
        ac.font = AVG_FONT; ac.fill = AVG_FILL
        ac.alignment = CENTER; ac.border = thin_border()

    data_range = f"B2:{get_column_letter(len(dates) + 1)}{len(tiers) + 1}"
    ws.conditional_formatting.add(data_range, colour_rule)
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False


def write_snr_sheet(wb, df):
    _pivot_sheet(wb, df, "SNR", "SNR by Tier", "SNR", "SNR", "Average SNR",
                 ColorScaleRule(start_type="min",  start_color="F8696B",
                                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                end_type="max",   end_color="63BE7B"))


def write_full_metrics_sheet(wb, df):
    ws = wb.create_sheet("Full Metrics")

    cols = ["Date", "Tier", "Features", "Total_Window_Anomalies",
            "Avg_In_Window", "Avg_Outside_Window", "SNR", "T_Stat", "P_Value"]
    col_widths = [14, 6, 42, 24, 16, 20, 10, 10, 10]

    ws.row_dimensions[1].height = 30
    for col_i, (col, width) in enumerate(zip(cols, col_widths), start=1):
        cell = ws.cell(row=1, column=col_i, value=col.replace("_", " "))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = width

    for row_i, (_, row) in enumerate(df[cols].iterrows(), start=2):
        fill = LIGHT_FILL if row_i % 2 == 0 else WHITE_FILL
        for col_i, col in enumerate(cols, start=1):
            val = row[col]
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = CENTER
            cell.border = thin_border()

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_sheet(input_path="output/comparison.csv", output_path="output/comparison.xlsx"):
    """Build a formatted Excel workbook from comparison.csv."""
    df = pd.read_csv(input_path)
    df["SNR"]     = pd.to_numeric(df["SNR"],     errors="coerce")
    df["T_Stat"]  = pd.to_numeric(df["T_Stat"],  errors="coerce")
    df["P_Value"] = pd.to_numeric(df["P_Value"], errors="coerce")
    df["Tier"]    = df["Tier"].astype(int)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_snr_sheet(wb, df)
    _pivot_sheet(wb, df, "T_Stat", "T-Statistic by Tier", "T-Stat", "T-Stat", "Average T-Stat",
                 ColorScaleRule(start_type="min",  start_color="F8696B",
                                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                end_type="max",   end_color="63BE7B"))
    _pivot_sheet(wb, df, "P_Value", "P-Value by Tier", "P-Value", "p-value", "Average p-value",
                 ColorScaleRule(start_type="min",  start_color="63BE7B",
                                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                end_type="max",   end_color="F8696B"))
    write_full_metrics_sheet(wb, df)
    wb.save(output_path)
    print(f"Saved {output_path}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input",  default="output/comparison.csv")
    p.add_argument("--output", default="output/comparison.xlsx")
    args = p.parse_args()
    build_sheet(args.input, args.output)


if __name__ == "__main__":
    main()
